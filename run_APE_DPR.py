def run_APE(data,data_path):
	from APE.alpha_S_to_M_without_Sigma import main_func

	#### switches
	set_start_sol_switch = 0
	callback_switch = 1



	results = main_func(data, \
		data_path, \
		set_start_sol_switch,\
		callback_switch,\
		Ns_sampled,\
		num_sampled_sens_points_ICP,\
		approx_sampled_model_points_APE,\
		approx_sampled_model_points_ICP,\
		TimeLimit_APE)
	return results



def run_DPR(data,data_path,corr):
	set_start_sol_switch = 0 
	callback_switch = 1 #### on or off

	from DPR.phi_with_Q import main_func  
	####### constants


	results = main_func(data, \
		data_path, \
		set_start_sol_switch,\
		callback_switch,\
		Ns_sampled,\
		num_sampled_sens_points_ICP,\
		approx_sampled_model_points_ICP,\
		model_tol_band,\
		TimeLimit_DPR,\
		corr )

	return results





def save_2_xl(a,b):
	total_buckets = max(len(a[0]),len(b[0]) )
	print a[0]
	print b[0]
	print total_buckets,"total_buckets"
	for i in range(total_buckets):
		if i >= len(a[0]):
			list_to_append = [ "NA" , str(b[0][i]),\
							   "NA" , str(b[1][i]),\
							   "NA" , str(b[2][i]),\
							   "NA" , str(b[3][i]),\
							   "NA" , str(b[4])]		


		elif i >= len(b[0]):	
			list_to_append = [str(a[0][i]), "NA" ,\
							  str(a[1][i]), "NA" ,\
							  str(a[2][i]), "NA" ,\
							  str(a[3][i]), "NA" ,\
							  str(a[4]), "NA" ]		

		else:	
			list_to_append = [str(a[0][i]), str(b[0][i]),\
							  str(a[1][i]), str(b[1][i]),\
							  str(a[2][i]), str(b[2][i]),\
							  str(a[3][i]), str(b[3][i]),\
							  str(a[4] ), str(b[4])]		

		ws1.append(list_to_append)  
		wb.save(file_name)


import os
import openpyxl


cwd = os.getcwd()
data_directory_name = "/"+ "bunny_noisy_varying_noise" 
data_path = cwd + data_directory_name +"/"
file_name = cwd+"/results/"+ data_directory_name + ".xlsx"

if 	os.path.exists(file_name):
	#### for loading the saved workbook
	print ("result file already exists")
else :
	#### for running dataset for the first time
	print ("result file does not exist")
	wb = openpyxl.Workbook(file_name)
	wb.save(file_name)

wb = openpyxl.load_workbook(file_name)
	



ws1 = wb.create_sheet("test APE")


####### parameters common to both
Ns_sampled = 30
num_sampled_sens_points_ICP = 500
approx_sampled_model_points_ICP = 500


####### APE specific 
approx_sampled_model_points_APE = 300
TimeLimit_APE = 300 #sec


######## DPR specific
TimeLimit_DPR = 300

	### parameters
# list_band_length = [5,20,30,50]
list_band_length = [20]
model_tol_band = list_band_length[0]

first_row =	["Ns_sampled", "num_sampled_sens_points_ICP","approx_sampled_model_points_ICP", "approx_sampled_model_points_APE",\
			 "TimeLimit_APE", "model_tol_band", "TimeLimit_DPR"]
ws1.append(first_row)
first_row =	[str(Ns_sampled), str(num_sampled_sens_points_ICP),str(approx_sampled_model_points_ICP), str(approx_sampled_model_points_APE),\
			 str(TimeLimit_APE), str(model_tol_band), str(TimeLimit_DPR)]
ws1.append(first_row)

wb.save(file_name)

for data in sorted(os.listdir(data_path)):
	
	print "data being used :", data

	ws1.append([data])
	ws1.append([data_path])
	third_row = ["APE_ang_err","DPR_ang_err", "APE_tra_err","DPR_tra_err","APE_tf","DPR_tf","APE_corr_matlab","DPR_corr_matlab","APE_obj_value","DPR_obj_val"] 
	ws1.append(third_row)
	wb.save(file_name)

	APE_ang_err_list,\
	APE_tra_err_list,\
	APE_tf_list,\
	APE_corr_python_list,\
	APE_corr_matlab_list,\
	APE_obj_val = run_APE(data,data_path)


	for i in range (len(list_band_length)):

		model_tol_band = list_band_length[i]
		first_row =	["Ns_sampled", "num_sampled_sens_points_ICP","approx_sampled_model_points_ICP", "approx_sampled_model_points_APE",\
				 "TimeLimit_APE", "model_tol_band", "TimeLimit_DPR"]
		ws1.append(first_row)
		first_row =	[str(Ns_sampled), str(num_sampled_sens_points_ICP),str(approx_sampled_model_points_ICP), str(approx_sampled_model_points_APE),\
					 str(TimeLimit_APE), str(model_tol_band), str(TimeLimit_DPR)]
		ws1.append(first_row)

		wb.save(file_name)
			
	 	DPR_ang_err_list,\
		DPR_tra_err_list,\
		DPR_tf_list,\
		DPR_corr_python_list,\
		DPR_corr_matlab_list,\
		DPR_obj_val = run_DPR(data,data_path,APE_corr_python_list[0])

		A = [APE_ang_err_list, APE_tra_err_list, APE_tf_list, APE_corr_matlab_list, APE_obj_val ]
		B = [DPR_ang_err_list, DPR_tra_err_list, DPR_tf_list, DPR_corr_matlab_list, DPR_obj_val ]
		save_2_xl(A,B)
		ws1.append([""])
		ws1.append([""])
